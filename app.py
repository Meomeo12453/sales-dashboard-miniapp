import streamlit as st
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import plotly.express as px
import colorsys
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Font

st.set_page_config(page_title="Sales Dashboard MiniApp", layout="wide")

st.title("📊 Sales Dashboard MiniApp")
st.markdown(
    "<small style='color:gray;'>Phiên bản tối ưu cho DABA Sài Gòn. Tải lên file Excel, xem bảng, biểu đồ, xuất file đẹp với màu nhóm tự động.</small>",
    unsafe_allow_html=True
)

# 1. Upload file
uploaded_file = st.file_uploader("Tải lên file Excel (xlsx)", type="xlsx")
if not uploaded_file:
    st.stop()

df = pd.read_excel(uploaded_file)
df['Mã khách hàng'] = df['Mã khách hàng'].astype(str)

# 2. Tính "Cấp dưới"
cap_duoi_list = []
for idx, row in df.iterrows():
    ma_kh = row['Mã khách hàng']
    ten_cap_tren, max_len = "", 0
    for idx2, row2 in df.iterrows():
        if idx == idx2: continue
        ma_cap_tren = row2['Mã khách hàng']
        if ma_cap_tren != ma_kh and ma_cap_tren in ma_kh:
            if len(ma_cap_tren) > max_len:
                ten_cap_tren = row2['Tên khách hàng']
                max_len = len(ma_cap_tren)
    cap_duoi_list.append(f"Cấp dưới {ten_cap_tren}" if ten_cap_tren else "")
df['Cấp dưới'] = cap_duoi_list

# 3. Tính "Số thuộc cấp"
so_thuoc_cap = []
for idx, row in df.iterrows():
    ma_kh = row['Mã khách hàng']
    count = sum((other_ma != ma_kh and other_ma.startswith(ma_kh)) for other_ma in df['Mã khách hàng'])
    so_thuoc_cap.append(count)
df['Số thuộc cấp'] = so_thuoc_cap

# 4. Doanh số hệ thống
def tinh_doanh_so_he_thong(df_in):
    dsht = []
    for idx, row in df_in.iterrows():
        ma_kh = row['Mã khách hàng']
        mask = (df_in['Mã khách hàng'] != ma_kh) & (df_in['Mã khách hàng'].str.startswith(ma_kh))
        subtotal = df_in.loc[mask, 'Tổng bán trừ trả hàng'].sum()
        dsht.append(subtotal)
    return dsht

df['Doanh số hệ thống'] = tinh_doanh_so_he_thong(df)

# 5. Hoa hồng
network = {
    'Catalyst':     {'comm_rate': 0.35, 'override_rate': 0.00},
    'Visionary':    {'comm_rate': 0.40, 'override_rate': 0.05},
    'Trailblazer':  {'comm_rate': 0.40, 'override_rate': 0.05},
}
df['comm_rate']     = df['Nhóm khách hàng'].map(lambda r: network.get(r, {}).get('comm_rate', 0))
df['override_rate'] = df['Nhóm khách hàng'].map(lambda r: network.get(r, {}).get('override_rate', 0))
df['override_comm'] = df['Doanh số hệ thống'] * df['override_rate']

st.success("✅ Đã xử lý xong dữ liệu. Xem bảng, biểu đồ hoặc tải kết quả ngay bên dưới.")

# 6. Hiển thị bảng dữ liệu
st.subheader("Bảng kết quả")
st.dataframe(df, use_container_width=True, hide_index=True)

# 7. Biểu đồ
tab1, tab2, tab3, tab4 = st.tabs(["Tổng bán & Hoa hồng hệ thống", "Sunburst", "Pareto", "Pie chart"])
with tab1:
    fig, ax = plt.subplots(figsize=(12,5))
    ind = np.arange(len(df))
    ax.bar(ind, df['Tổng bán trừ trả hàng'], width=0.5, label='Tổng bán cá nhân')
    ax.bar(ind, df['override_comm'], width=0.5, bottom=df['Tổng bán trừ trả hàng'], label='Hoa hồng hệ thống')
    ax.set_ylabel('Số tiền (VND)')
    ax.set_title('Tổng bán & Hoa hồng hệ thống từng cá nhân')
    ax.set_xticks(ind)
    ax.set_xticklabels(df['Tên khách hàng'], rotation=60, ha='right')
    ax.legend()
    st.pyplot(fig)

with tab2:
    fig2 = px.sunburst(
        df,
        path=['Nhóm khách hàng', 'Tên khách hàng'],
        values='Tổng bán trừ trả hàng',
        title="Sơ đồ hệ thống cấp bậc & doanh số"
    )
    st.plotly_chart(fig2, use_container_width=True)

with tab3:
    df_sorted = df.sort_values('Tổng bán trừ trả hàng', ascending=False)
    cum_sum = df_sorted['Tổng bán trừ trả hàng'].cumsum()
    cum_perc = 100 * cum_sum / df_sorted['Tổng bán trừ trả hàng'].sum()
    fig3, ax1 = plt.subplots(figsize=(10,5))
    ax1.bar(np.arange(len(df_sorted)), df_sorted['Tổng bán trừ trả hàng'], label="Doanh số")
    ax1.set_ylabel('Doanh số')
    ax1.set_xticks(range(len(df_sorted)))
    ax1.set_xticklabels(df_sorted['Tên khách hàng'], rotation=60, ha='right')
    ax2 = ax1.twinx()
    ax2.plot(np.arange(len(df_sorted)), cum_perc, color='red', marker='o', label='Tích lũy (%)')
    ax2.set_ylabel('Tỷ lệ tích lũy (%)')
    ax1.set_title('Biểu đồ Pareto: Doanh số & tỷ trọng tích lũy')
    fig3.tight_layout()
    st.pyplot(fig3)

with tab4:
    fig4, ax4 = plt.subplots(figsize=(6,6))
    s = df.groupby('Nhóm khách hàng')['Tổng bán trừ trả hàng'].sum()
    ax4.pie(s, labels=s.index, autopct='%1.1f%%')
    ax4.set_title('Tỷ trọng doanh số theo nhóm khách hàng')
    st.pyplot(fig4)

# 8. Xuất file đẹp, màu nhóm tương phản, cho tải về
output_file = 'sales_report_dep.xlsx'
df.to_excel(output_file, index=False)

# --- Định dạng file ---
from openpyxl import load_workbook
wb = load_workbook(output_file)
ws = wb.active

header_fill = PatternFill(start_color='FFE699', end_color='FFE699', fill_type='solid')
header_font = Font(bold=True, color='000000')
header_align = Alignment(horizontal='center', vertical='center')
for col in range(1, ws.max_column + 1):
    cell = ws.cell(row=1, column=col)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_align

money_keywords = ['bán', 'doanh số', 'tiền', 'hoa hồng', 'comm', 'VND']
cols_money = [col[0].column for col in ws.iter_cols(1, ws.max_column)
              if any(key in (col[0].value or '').lower() for key in money_keywords)]

col_makh = [cell.value for cell in ws[1]].index('Mã khách hàng')+1
col_role = [cell.value for cell in ws[1]].index('Nhóm khách hàng')+1

# Gom nhóm mã KH: tìm prefix dài nhất có nhiều dòng nhất
all_codes = [str(ws.cell(row=i, column=col_makh).value) for i in range(2, ws.max_row+1)]
prefix_groups = {}
for length in range(len(max(all_codes, key=len)), 0, -1):
    prefix_count = {}
    for code in all_codes:
        if len(code) < length:
            continue
        prefix = code[:length]
        prefix_count.setdefault(prefix, []).append(code)
    for prefix, codes in prefix_count.items():
        if len(codes) > 1:
            prefix_groups[prefix] = codes

row_to_prefix = {}
for idx, code in enumerate(all_codes):
    best_prefix = ''
    best_len = 0
    for prefix in prefix_groups.keys():
        if code.startswith(prefix) and len(prefix) > best_len:
            best_prefix = prefix
            best_len = len(prefix)
    row_to_prefix[idx+2] = best_prefix if best_prefix else code

prefix_set = set(row_to_prefix.values())
prefix_list = sorted(prefix_set)
def get_contrasting_color(idx, total):
    h = idx / total
    r, g, b = colorsys.hsv_to_rgb(h, 0.65, 1)
    return "%02X%02X%02X" % (int(r*255), int(g*255), int(b*255))
prefix_to_color = {prefix: PatternFill(start_color=get_contrasting_color(i, len(prefix_list)),
                                       end_color=get_contrasting_color(i, len(prefix_list)),
                                       fill_type='solid')
                   for i, prefix in enumerate(prefix_list)}

for row in range(2, ws.max_row + 1):
    role = ws.cell(row=row, column=col_role).value
    if role == 'Trailblazer':
        fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    else:
        fill = prefix_to_color[row_to_prefix[row]]
    for col in range(1, ws.max_column + 1):
        ws.cell(row=row, column=col).fill = fill

for col in range(1, ws.max_column + 1):
    for row in range(2, ws.max_row+1):
        cell = ws.cell(row=row, column=col)
        if col in cols_money:
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0'
            cell.alignment = Alignment(horizontal='right', vertical='center')
        else:
            cell.alignment = Alignment(horizontal='center', vertical='center')

for col in ws.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        val = str(cell.value) if cell.value else ""
        max_length = max(max_length, len(val.encode('utf8'))//2+2)
    ws.column_dimensions[column].width = max(10, min(40, max_length))

# Xuất file ra memory để người dùng tải về
bio = BytesIO()
wb.save(bio)
st.download_button(
    label="📥 Tải file Excel đã định dạng",
    data=bio.getvalue(),
    file_name=output_file,
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
